import os
import csv
from openpyxl import load_workbook

dir_path = os.path.dirname(os.path.realpath(__file__))
#dir_path = "scripts"

wb = load_workbook(dir_path + "/../data/UBS_PricesAndEarnings_OpenData.xlsx")
ws = wb.active
#print(tuple(ws.columns))

def get_first_2018_index():
  for i in range(1,200000):
    if (ws['A'+str(i)].value == 2018):
      print(i)
      break

interesting_labels = [
  'Earnings: Average annual (gross)',
  'General: Domestic Purchasing Power',
  'Prices: Housing',
  'Prices: Total expenditure on goods and services',
  'Prices: Food'
]

# A - year
# B - city
# C - data type
# D - value description
# E - value

def print_next_five(base):
    for i in range(0,4):
        print("{} {} {} {} {}".format(ws['A' + str(base + i)].value,
                                      ws['B' + str(base + i)].value,
                                      ws['C' + str(base + i)].value,
                                      ws['D' + str(base + i)].value,
                                      ws['E' + str(base + i)].value))


# label -> list of tuples of size 3:
ubs_data = {}
for label in interesting_labels:
  ubs_data[label] = []

for label in interesting_labels:
  for i in range(1,200000):
    if ws['A' + str(i)].value == 2018 and \
       ws['C' + str(i)].value == label and \
       (ws['D' + str(i)].value == '(USD)' or (label == 'General: Domestic Purchasing Power' and ws['D' + str(i)].value == 'Net (New York = 100)') ):
        ubs_data[label].append((ws['B' + str(i)].value, ws['D' + str(i)].value, ws['E' + str(i)].value))

with open('ubs_data.csv', 'w', newline='') as csvfile:
    spamwriter = csv.writer(csvfile, delimiter=',')
    spamwriter.writerow(['City', 'Label', 'Value description', 'Value'])
    for label in interesting_labels:
      for d in ubs_data[label]:  
        spamwriter.writerow([d[0], label, d[1], d[2]])

